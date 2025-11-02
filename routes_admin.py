from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file, make_response
from flask_login import login_required, current_user
from models import db, User, Role, LeaveRequest, LeaveType, Schedule, Attendance, SystemSettings, Notification
from datetime import datetime, timedelta
from werkzeug.utils import secure_filename
import openpyxl
from openpyxl import Workbook
import os
from io import BytesIO
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.enums import TA_RIGHT, TA_CENTER

admin_bp = Blueprint('admin', __name__, url_prefix='/admin')

# التحقق من صلاحيات الإدارة
def admin_required():
    return current_user.is_authenticated and current_user.role in [Role.MAIN_ADMIN, Role.SUB_ADMIN]

# لوحة تحكم الإدارة
@admin_bp.route('/dashboard')
@login_required
def dashboard():
    if not admin_required():
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('index'))
    
    # إحصائيات عامة
    total_employees = User.query.filter_by(role=Role.EMPLOYEE).count()
    total_supervisors = User.query.filter(
        User.role.in_([Role.MAIN_SUPERVISOR, Role.SUB_SUPERVISOR])
    ).count()
    pending_leaves = LeaveRequest.query.filter_by(status='قيد الانتظار').count()
    
    # الحضور اليوم
    today = datetime.now().date()
    present_today = Attendance.query.filter(
        Attendance.date == today,
        Attendance.status == 'حاضر'
    ).count()
    
    # آخر طلبات الإجازات
    recent_leaves = LeaveRequest.query.order_by(LeaveRequest.created_at.desc()).limit(10).all()
    
    return render_template('admin/dashboard.html',
                         total_employees=total_employees,
                         total_supervisors=total_supervisors,
                         pending_leaves=pending_leaves,
                         present_today=present_today,
                         recent_leaves=recent_leaves)

# إدارة المشرفين
@admin_bp.route('/supervisors')
@login_required
def supervisors():
    if not admin_required():
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('index'))
    
    supervisors_list = User.query.filter(
        User.role.in_([Role.MAIN_SUPERVISOR, Role.SUB_SUPERVISOR])
    ).all()
    
    return render_template('admin/supervisors.html', supervisors=supervisors_list)

# إضافة مشرف جديد
@admin_bp.route('/supervisors/add', methods=['GET', 'POST'])
@login_required
def add_supervisor():
    if not admin_required():
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        national_id = request.form.get('national_id')
        name = request.form.get('name')
        password = request.form.get('password')
        role = request.form.get('role')
        gender = request.form.get('gender')
        department = request.form.get('department')
        shift_start = request.form.get('shift_start')
        shift_end = request.form.get('shift_end')
        shift_time = f"{shift_start} - {shift_end}" if shift_start and shift_end else None
        
        # التحقق من عدم وجود الهوية مسبقاً
        existing = User.query.filter_by(national_id=national_id).first()
        if existing:
            flash('رقم الهوية موجود مسبقاً', 'danger')
            return redirect(url_for('admin.add_supervisor'))
        
        supervisor = User(
            national_id=national_id,
            name=name,
            role=role,
            gender=gender,
            department=department,
            shift_time=shift_time
        )
        supervisor.set_password(password)
        
        db.session.add(supervisor)
        db.session.commit()
        
        flash('تم إضافة المشرف بنجاح', 'success')
        return redirect(url_for('admin.supervisors'))
    
    return render_template('admin/add_supervisor.html')

# إدارة الموظفين
@admin_bp.route('/employees')
@login_required
def employees():
    if not admin_required():
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('index'))
    
    employees_list = User.query.filter_by(role=Role.EMPLOYEE).all()
    return render_template('admin/employees.html', employees=employees_list)

# إضافة موظف يدوياً
@admin_bp.route('/employees/add', methods=['GET', 'POST'])
@login_required
def add_employee():
    if not admin_required():
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        national_id = request.form.get('national_id')
        name = request.form.get('name')
        gender = request.form.get('gender')
        department = request.form.get('department')
        
        # التحقق من عدم وجود الهوية
        existing = User.query.filter_by(national_id=national_id).first()
        if existing:
            flash('رقم الهوية موجود مسبقاً', 'danger')
            return redirect(url_for('admin.add_employee'))
        
        employee = User(
            national_id=national_id,
            name=name,
            role=Role.EMPLOYEE,
            gender=gender,
            department=department
        )
        # كلمة مرور افتراضية = رقم الهوية
        employee.set_password(national_id)
        
        db.session.add(employee)
        db.session.commit()
        
        flash('تم إضافة الموظف بنجاح', 'success')
        return redirect(url_for('admin.employees'))
    
    return render_template('admin/add_employee.html')

# تعديل موظف
@admin_bp.route('/employees/edit/<int:employee_id>', methods=['GET', 'POST'])
@login_required
def edit_employee(employee_id):
    if not admin_required():
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('index'))
    
    employee = User.query.get_or_404(employee_id)
    
    if request.method == 'POST':
        employee.name = request.form.get('name')
        employee.national_id = request.form.get('national_id')
        employee.gender = request.form.get('gender')
        employee.department = request.form.get('department')
        
        shift_start = request.form.get('shift_start')
        shift_end = request.form.get('shift_end')
        if shift_start and shift_end:
            employee.shift_time = f"{shift_start} - {shift_end}"
        
        password = request.form.get('password')
        if password:
            employee.set_password(password)
        
        db.session.commit()
        flash('تم تعديل الموظف بنجاح', 'success')
        return redirect(url_for('admin.employees'))
    
    # تحويل shift_time إلى start و end
    shift_parts = employee.shift_time.split(' - ') if employee.shift_time else ['', '']
    
    return render_template('admin/edit_employee.html', 
                          employee=employee,
                          shift_start=shift_parts[0] if len(shift_parts) > 0 else '',
                          shift_end=shift_parts[1] if len(shift_parts) > 1 else '')

# رفع ملف Excel للموظفين
@admin_bp.route('/employees/upload', methods=['GET', 'POST'])
@login_required
def upload_employees():
    if not admin_required():
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('لم يتم اختيار ملف', 'danger')
            return redirect(url_for('admin.upload_employees'))
        
        file = request.files['file']
        if file.filename == '':
            flash('لم يتم اختيار ملف', 'danger')
            return redirect(url_for('admin.upload_employees'))
        
        if file and file.filename.endswith(('.xlsx', '.xls')):
            try:
                wb = openpyxl.load_workbook(file)
                sheet = wb.active
                
                added_count = 0
                skipped_count = 0
                
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    # التحقق من وجود بيانات
                    if not row or not row[0]:
                        continue
                    
                    # قراءة البيانات بالترتيب الصحيح (حسب الصورة)
                    name = str(row[0]).strip() if row[0] else ''  # الاسم
                    national_id = str(row[1]).strip() if row[1] else ''  # الهوية
                    period = str(row[2]).strip() if row[2] else ''  # الفترة
                    work_time = str(row[3]).strip() if row[3] else ''  # الوقت
                    rest_days = str(row[4]).strip() if row[4] else ''  # أيام الراحة
                    department = str(row[5]).strip() if len(row) > 5 and row[5] else 'الحلقات'  # القسم
                    gender = str(row[6]).strip() if len(row) > 6 and row[6] else 'ذكر'  # الجنس
                    
                    # التحقق من وجود الموظف
                    existing = User.query.filter_by(national_id=national_id).first()
                    if existing:
                        # تحديث معلومات الجدول للموظف الموجود
                        existing.period = period
                        existing.work_time = work_time
                        existing.rest_days = rest_days
                        existing.department = department
                        existing.gender = gender
                        skipped_count += 1
                        continue
                    
                    employee = User(
                        national_id=national_id,
                        name=name,
                        role=Role.EMPLOYEE,
                        gender=gender,
                        department=department,
                        period=period,
                        work_time=work_time,
                        rest_days=rest_days
                    )
                    employee.set_password(national_id)
                    
                    db.session.add(employee)
                    added_count += 1
                
                db.session.commit()
                flash(f'تم إضافة {added_count} موظف. تم تخطي {skipped_count} موظف موجود مسبقاً', 'success')
                
            except Exception as e:
                flash(f'حدث خطأ أثناء قراءة الملف: {str(e)}', 'danger')
        else:
            flash('نوع الملف غير مدعوم. الرجاء رفع ملف Excel', 'danger')
        
        return redirect(url_for('admin.employees'))
    
    return render_template('admin/upload_employees.html')

# تحميل نموذج Excel
@admin_bp.route('/employees/download-template')
@login_required
def download_template():
    if not admin_required():
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('index'))
    
    wb = Workbook()
    sheet = wb.active
    sheet.title = "الموظفين"
    
    # العناوين بالترتيب الصحيح
    headers = ['الاسم', 'الهوية', 'الفترة', 'الوقت', 'الراحة', 'القسم', 'الجنس']
    sheet.append(headers)
    
    # حفظ في الذاكرة
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='نموذج_الموظفين.xlsx'
    )

# إسناد الموظفين للمشرفين
@admin_bp.route('/assign-employees', methods=['GET', 'POST'])
@login_required
def assign_employees():
    if not admin_required():
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('index'))
    
    supervisors = User.query.filter(
        User.role.in_([Role.MAIN_SUPERVISOR, Role.SUB_SUPERVISOR])
    ).all()
    
    employees = User.query.filter_by(role=Role.EMPLOYEE).all()
    
    if request.method == 'POST':
        supervisor_id = request.form.get('supervisor_id', type=int)
        employee_ids = request.form.getlist('employee_ids')
        
        for emp_id in employee_ids:
            employee = User.query.get(int(emp_id))
            if employee:
                employee.supervisor_id = supervisor_id
        
        db.session.commit()
        flash('تم إسناد الموظفين بنجاح', 'success')
        return redirect(url_for('admin.assign_employees'))
    
    return render_template('admin/assign_employees.html', 
                         supervisors=supervisors,
                         employees=employees)

# إدارة أنواع الإجازات
@admin_bp.route('/leave-types')
@login_required
def leave_types():
    if not admin_required():
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('index'))
    
    leave_types_list = LeaveType.query.all()
    return render_template('admin/leave_types.html', leave_types=leave_types_list)

# إضافة نوع إجازة
@admin_bp.route('/leave-types/add', methods=['POST'])
@login_required
def add_leave_type():
    if not admin_required():
        return jsonify({'success': False}), 403
    
    name = request.form.get('name')
    max_days = request.form.get('max_days', type=int)
    requires_attachment = request.form.get('requires_attachment') == 'on'
    
    leave_type = LeaveType(
        name=name,
        max_days=max_days,
        requires_attachment=requires_attachment
    )
    
    db.session.add(leave_type)
    db.session.commit()
    
    flash('تم إضافة نوع الإجازة بنجاح', 'success')
    return redirect(url_for('admin.leave_types'))

# تعديل نوع إجازة
@admin_bp.route('/leave-types/edit/<int:leave_type_id>', methods=['POST'])
@login_required
def edit_leave_type(leave_type_id):
    if not admin_required():
        return jsonify({'success': False}), 403
    
    leave_type = LeaveType.query.get_or_404(leave_type_id)
    
    leave_type.name = request.form.get('name')
    leave_type.max_days = request.form.get('max_days', type=int)
    leave_type.requires_attachment = request.form.get('requires_attachment') == 'on'
    leave_type.is_active = request.form.get('is_active') == 'on'
    
    db.session.commit()
    
    flash('تم تعديل نوع الإجازة بنجاح', 'success')
    return redirect(url_for('admin.leave_types'))

# حذف نوع إجازة
@admin_bp.route('/leave-types/delete/<int:leave_type_id>', methods=['POST'])
@login_required
def delete_leave_type(leave_type_id):
    if not admin_required():
        return jsonify({'success': False}), 403
    
    leave_type = LeaveType.query.get_or_404(leave_type_id)
    
    # التحقق من عدم وجود طلبات إجازة مرتبطة بهذا النوع
    if leave_type.leave_requests.count() > 0:
        flash('لا يمكن حذف نوع الإجازة لأنه مرتبط بطلبات إجازة', 'danger')
        return redirect(url_for('admin.leave_types'))
    
    db.session.delete(leave_type)
    db.session.commit()
    
    flash('تم حذف نوع الإجازة بنجاح', 'success')
    return redirect(url_for('admin.leave_types'))

# إعدادات النظام
@admin_bp.route('/settings', methods=['GET', 'POST'])
@login_required
def settings():
    if current_user.role != Role.MAIN_ADMIN:
        flash('هذه الصفحة متاحة فقط لمدير النظام الأساسي', 'danger')
        return redirect(url_for('index'))
    
    system_settings = SystemSettings.query.first()
    if not system_settings:
        system_settings = SystemSettings()
        db.session.add(system_settings)
        db.session.commit()
    
    if request.method == 'POST':
        system_settings.system_name = request.form.get('system_name')
        system_settings.primary_color = request.form.get('primary_color')
        system_settings.secondary_color = request.form.get('secondary_color')
        system_settings.accent_color = request.form.get('accent_color')
        system_settings.attachment_retention_days = request.form.get('attachment_retention_days', type=int)
        
        db.session.commit()
        flash('تم تحديث الإعدادات بنجاح', 'success')
        return redirect(url_for('admin.settings'))
    
    return render_template('admin/settings.html', settings=system_settings)

# التقارير
@admin_bp.route('/reports')
@login_required
def reports():
    if not admin_required():
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('index'))
    
    return render_template('admin/reports.html')

# تقرير الإجازات
@admin_bp.route('/reports/leaves')
@login_required
def report_leaves():
    if not admin_required():
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('index'))
    
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    status = request.args.get('status')
    
    query = LeaveRequest.query
    
    if start_date:
        query = query.filter(LeaveRequest.start_date >= datetime.strptime(start_date, '%Y-%m-%d').date())
    if end_date:
        query = query.filter(LeaveRequest.end_date <= datetime.strptime(end_date, '%Y-%m-%d').date())
    if status:
        query = query.filter(LeaveRequest.status == status)
    
    leaves = query.order_by(LeaveRequest.created_at.desc()).all()
    
    return render_template('admin/report_leaves.html', leaves=leaves)

# تقرير الحضور
@admin_bp.route('/reports/attendance')
@login_required
def report_attendance():
    if not admin_required():
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('index'))
    
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    employee_id = request.args.get('employee_id', type=int)
    
    query = Attendance.query
    
    if start_date:
        query = query.filter(Attendance.date >= datetime.strptime(start_date, '%Y-%m-%d').date())
    if end_date:
        query = query.filter(Attendance.date <= datetime.strptime(end_date, '%Y-%m-%d').date())
    if employee_id:
        query = query.filter(Attendance.employee_id == employee_id)
    
    records = query.order_by(Attendance.date.desc()).all()
    employees = User.query.filter_by(role=Role.EMPLOYEE).all()
    
    return render_template('admin/report_attendance.html', records=records, employees=employees)

# تخصيص المظهر (الألوان والشعار)
@admin_bp.route('/customize', methods=['GET', 'POST'])
@login_required
def customize():
    if current_user.role != Role.MAIN_ADMIN:
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('index'))
    
    settings = SystemSettings.query.first()
    if not settings:
        settings = SystemSettings()
        db.session.add(settings)
        db.session.commit()
    
    if request.method == 'POST':
        # تحديث الألوان
        settings.primary_color = request.form.get('primary_color', '#0d7377')
        settings.secondary_color = request.form.get('secondary_color', '#14FFEC')
        settings.accent_color = request.form.get('accent_color', '#323232')
        settings.system_name = request.form.get('system_name', 'نظام إدارة معلمي الحلقات - مكة المكرمة')
        
        # رفع الشعار
        if 'logo' in request.files:
            logo = request.files['logo']
            if logo and logo.filename:
                filename = secure_filename('logo_' + logo.filename)
                logo_folder = os.path.join('static', 'images')
                os.makedirs(logo_folder, exist_ok=True)
                logo_path = os.path.join(logo_folder, filename)
                logo.save(logo_path)
                settings.logo_path = filename  # نحفظ اسم الملف فقط
        
        db.session.commit()
        flash('تم تحديث التخصيصات بنجاح', 'success')
        return redirect(url_for('admin.customize'))
    
    return render_template('admin/customize.html', settings=settings)

# طباعة تقرير الإجازات PDF
@admin_bp.route('/reports/leaves/pdf')
@login_required
def report_leaves_pdf():
    if not admin_required():
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('index'))
    
    # جلب البيانات
    employee_id = request.args.get('employee_id', type=int)
    leave_type_id = request.args.get('leave_type_id', type=int)
    status = request.args.get('status')
    
    # تحديد الـ join بشكل صريح لتجنب AmbiguousForeignKeysError
    query = LeaveRequest.query.join(User, LeaveRequest.employee_id == User.id).join(LeaveType)
    
    if employee_id:
        query = query.filter(LeaveRequest.employee_id == employee_id)
    if leave_type_id:
        query = query.filter(LeaveRequest.leave_type_id == leave_type_id)
    if status:
        query = query.filter(LeaveRequest.status == status)
    
    leaves = query.order_by(LeaveRequest.created_at.desc()).all()
    
    # إنشاء PDF مع دعم العربية
    from reportlab.pdfbase.pdfmetrics import registerFont
    from reportlab.pdfbase.ttfonts import TTFont
    from arabic_reshaper import reshape
    from bidi.algorithm import get_display
    
    # تسجيل خط عربي (استخدام Arial Unicode MS)
    try:
        registerFont(TTFont('Arabic', 'C:/Windows/Fonts/arial.ttf'))
        arabic_font = 'Arabic'
    except:
        # إذا فشل، نستخدم الخط الافتراضي
        arabic_font = 'Helvetica'
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    
    elements = []
    
    # العنوان بالعربية
    title_text = 'تقرير الإجازات'
    reshaped_title = reshape(title_text)
    bidi_title = get_display(reshaped_title)
    
    title_style = ParagraphStyle(
        'TitleStyle',
        fontSize=18,
        alignment=TA_CENTER,
        spaceAfter=20,
        fontName=arabic_font
    )
    elements.append(Paragraph(bidi_title, title_style))
    elements.append(Spacer(1, 20))
    
    # الجدول
    def arabic_text(text):
        """تحويل النص العربي للعرض الصحيح"""
        reshaped = reshape(text)
        return get_display(reshaped)
    
    headers = ['الموظف', 'نوع الإجازة', 'من تاريخ', 'إلى تاريخ', 'الأيام', 'الحالة']
    data = [[arabic_text(h) for h in headers]]
    
    for leave in leaves:
        data.append([
            arabic_text(leave.employee.name),
            arabic_text(leave.leave_type.name),
            leave.start_date.strftime('%Y-%m-%d'),
            leave.end_date.strftime('%Y-%m-%d'),
            str(leave.days_count),
            arabic_text(leave.status)
        ])
    
    table = Table(data, colWidths=[4*cm, 3*cm, 2.5*cm, 2.5*cm, 2*cm, 2.5*cm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0d7377')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, -1), arabic_font),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name=f'leaves_report_{datetime.now().strftime("%Y%m%d")}.pdf', mimetype='application/pdf')

# طباعة تقرير الحضور PDF
@admin_bp.route('/reports/attendance/pdf')
@login_required  
def report_attendance_pdf():
    if not admin_required():
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('index'))
    
    # جلب البيانات
    employee_id = request.args.get('employee_id', type=int)
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    query = Attendance.query.join(User, Attendance.employee_id == User.id)
    
    if start_date:
        query = query.filter(Attendance.date >= datetime.strptime(start_date, '%Y-%m-%d').date())
    if end_date:
        query = query.filter(Attendance.date <= datetime.strptime(end_date, '%Y-%m-%d').date())
    if employee_id:
        query = query.filter(Attendance.employee_id == employee_id)
    
    records = query.order_by(Attendance.date.desc()).all()
    
    # إنشاء PDF مع دعم العربية
    from reportlab.pdfbase.pdfmetrics import registerFont
    from reportlab.pdfbase.ttfonts import TTFont
    from arabic_reshaper import reshape
    from bidi.algorithm import get_display
    
    # تسجيل خط عربي
    try:
        registerFont(TTFont('Arabic', 'C:/Windows/Fonts/arial.ttf'))
        arabic_font = 'Arabic'
    except:
        arabic_font = 'Helvetica'
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    
    elements = []
    
    # العنوان بالعربية
    title_text = 'تقرير الحضور والغياب'
    reshaped_title = reshape(title_text)
    bidi_title = get_display(reshaped_title)
    
    title_style = ParagraphStyle(
        'TitleStyle',
        fontSize=18,
        alignment=TA_CENTER,
        spaceAfter=20,
        fontName=arabic_font
    )
    elements.append(Paragraph(bidi_title, title_style))
    elements.append(Spacer(1, 20))
    
    # الجدول
    def arabic_text(text):
        """تحويل النص العربي للعرض الصحيح"""
        reshaped = reshape(text)
        return get_display(reshaped)
    
    headers = ['الموظف', 'التاريخ', 'الحالة', 'الملاحظات']
    data = [[arabic_text(h) for h in headers]]
    
    for record in records:
        data.append([
            arabic_text(record.employee.name),
            record.date.strftime('%Y-%m-%d'),
            arabic_text(record.status),
            arabic_text(record.notes) if record.notes else '-'
        ])
    
    table = Table(data, colWidths=[5*cm, 3*cm, 3*cm, 5*cm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0d7377')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, -1), arabic_font),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name=f'attendance_report_{datetime.now().strftime("%Y%m%d")}.pdf', mimetype='application/pdf')

# حذف بيانات الاختبار (مدير أساسي فقط)
@admin_bp.route('/delete-test-data', methods=['POST'])
@login_required
def delete_test_data_route():
    if current_user.role != Role.MAIN_ADMIN:
        flash('ليس لديك صلاحية لهذه العملية', 'danger')
        return redirect(url_for('admin.dashboard'))
    
    try:
        # حذف سجلات الحضور للموظفين التجريبيين
        test_employees = User.query.filter(
            User.national_id.like('4000%') | User.national_id.like('5000%')
        ).all()
        
        for emp in test_employees:
            Attendance.query.filter_by(employee_id=emp.id).delete()
            LeaveRequest.query.filter_by(employee_id=emp.id).delete()
            Schedule.query.filter_by(employee_id=emp.id).delete()
        
        # حذف المستخدمين التجريبيين
        User.query.filter(
            User.national_id.like('4000%') | User.national_id.like('5000%')
        ).delete(synchronize_session=False)
        
        User.query.filter(
            User.national_id.like('2000%') | User.national_id.like('3000%')
        ).delete(synchronize_session=False)
        
        db.session.commit()
        flash('تم حذف بيانات الاختبار بنجاح', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ: {str(e)}', 'danger')
    
    return redirect(url_for('admin.dashboard'))

# عرض طلبات الإجازات للإدارة
@admin_bp.route('/leave-requests')
@login_required
def leave_requests():
    if not admin_required():
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('index'))
    
    # جلب جميع طلبات الإجازات
    status_filter = request.args.get('status', 'قيد الانتظار')
    
    query = LeaveRequest.query.join(User, LeaveRequest.employee_id == User.id).join(LeaveType)
    
    if status_filter and status_filter != 'all':
        query = query.filter(LeaveRequest.status == status_filter)
    
    requests = query.order_by(LeaveRequest.created_at.desc()).all()
    
    return render_template('admin/leave_requests.html', requests=requests, status_filter=status_filter)

# مراجعة طلب إجازة (قبول/رفض)
@admin_bp.route('/review-leave/<int:request_id>', methods=['POST'])
@login_required
def review_leave(request_id):
    if not admin_required():
        flash('ليس لديك صلاحية لهذه العملية', 'danger')
        return redirect(url_for('index'))
    
    leave_request = LeaveRequest.query.get_or_404(request_id)
    action = request.form.get('action')
    notes = request.form.get('notes', '')
    
    if action == 'approve':
        leave_request.status = 'مقبول'
        leave_request.reviewed_by = current_user.id
        leave_request.reviewed_at = datetime.now()
        leave_request.review_notes = notes
        flash('تم قبول طلب الإجازة بنجاح', 'success')
    elif action == 'reject':
        leave_request.status = 'مرفوض'
        leave_request.reviewed_by = current_user.id
        leave_request.reviewed_at = datetime.now()
        leave_request.review_notes = notes
        flash('تم رفض طلب الإجازة', 'warning')
    
    db.session.commit()
    
    return redirect(url_for('admin.leave_requests'))

# عرض جدول الحلقات
@admin_bp.route('/schedules-table')
@login_required
def schedules_table():
    if not admin_required():
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('index'))
    
    # جلب جميع المعلمين (الموظفين) مع معلومات جدولهم
    employees = User.query.filter_by(role=Role.EMPLOYEE, is_active=True).order_by(User.name).all()
    
    return render_template('admin/schedules_table.html', employees=employees)

# حذف جميع بيانات المعلمين
@admin_bp.route('/delete-all-employees', methods=['POST'])
@login_required
def delete_all_employees():
    if current_user.role != Role.MAIN_ADMIN:
        flash('ليس لديك صلاحية لهذه العملية', 'danger')
        return redirect(url_for('admin.schedules_table'))
    
    try:
        # حذف جميع سجلات المعلمين
        employees = User.query.filter_by(role=Role.EMPLOYEE).all()
        
        for emp in employees:
            # حذف السجلات المرتبطة
            Attendance.query.filter_by(employee_id=emp.id).delete()
            LeaveRequest.query.filter_by(employee_id=emp.id).delete()
            Schedule.query.filter_by(employee_id=emp.id).delete()
            
            # حذف الإشعارات المرتبطة
            from models import Notification
            Notification.query.filter_by(user_id=emp.id).delete()
            
            # حذف الموظف
            db.session.delete(emp)
        
        db.session.commit()
        flash(f'تم حذف جميع بيانات المعلمين بنجاح ({len(employees)} معلم)', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ: {str(e)}', 'danger')
    
    return redirect(url_for('admin.schedules_table'))

# صفحة إعدادات الحساب
@admin_bp.route('/account-settings', methods=['GET', 'POST'])
@login_required
def account_settings():
    if not admin_required():
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        action = request.form.get('action')
        
        # تغيير كلمة السر
        if action == 'change_password':
            current_password = request.form.get('current_password')
            new_password = request.form.get('new_password')
            confirm_password = request.form.get('confirm_password')
            
            # التحقق من كلمة السر الحالية
            if not current_user.check_password(current_password):
                flash('كلمة السر الحالية غير صحيحة', 'danger')
                return redirect(url_for('admin.account_settings'))
            
            # التحقق من تطابق كلمة السر الجديدة
            if new_password != confirm_password:
                flash('كلمة السر الجديدة غير متطابقة', 'danger')
                return redirect(url_for('admin.account_settings'))
            
            # التحقق من طول كلمة السر
            if len(new_password) < 6:
                flash('كلمة السر يجب أن تكون 6 أحرف على الأقل', 'danger')
                return redirect(url_for('admin.account_settings'))
            
            # تغيير كلمة السر
            current_user.set_password(new_password)
            db.session.commit()
            flash('تم تغيير كلمة السر بنجاح', 'success')
            return redirect(url_for('admin.account_settings'))
        
        # تغيير رقم الهوية
        elif action == 'change_national_id':
            new_national_id = request.form.get('new_national_id')
            password_confirm = request.form.get('password_confirm')
            
            # التحقق من كلمة السر
            if not current_user.check_password(password_confirm):
                flash('كلمة السر غير صحيحة', 'danger')
                return redirect(url_for('admin.account_settings'))
            
            # التحقق من صحة رقم الهوية
            if len(new_national_id) != 10 or not new_national_id.isdigit():
                flash('رقم الهوية يجب أن يكون 10 أرقام', 'danger')
                return redirect(url_for('admin.account_settings'))
            
            # التحقق من عدم تكرار رقم الهوية
            existing = User.query.filter_by(national_id=new_national_id).first()
            if existing and existing.id != current_user.id:
                flash('رقم الهوية موجود لمستخدم آخر', 'danger')
                return redirect(url_for('admin.account_settings'))
            
            # تغيير رقم الهوية
            old_id = current_user.national_id
            current_user.national_id = new_national_id
            db.session.commit()
            flash(f'تم تغيير رقم الهوية من {old_id} إلى {new_national_id}', 'success')
            return redirect(url_for('admin.account_settings'))
    
    return render_template('admin/account_settings.html')
